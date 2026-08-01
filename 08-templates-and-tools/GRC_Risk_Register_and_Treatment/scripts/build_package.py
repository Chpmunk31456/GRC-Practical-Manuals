#!/usr/bin/env python3
from __future__ import annotations

import csv
import hashlib
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "publication"
CHAPTERS = [
    ROOT / "chapters/01_Risk_Management_Foundations.md",
    ROOT / "chapters/02_Risk_Identification_and_Scenario_Writing.md",
    ROOT / "chapters/03_Risk_Analysis_and_Scoring.md",
    ROOT / "chapters/04_Risk_Response_Treatment_and_Acceptance.md",
    ROOT / "chapters/05_Monitoring_Reporting_and_Escalation.md",
    ROOT / "chapters/06_Operating_Procedure_and_Worked_Example.md",
]


def assemble_master() -> Path:
    template = (ROOT / "English_Master.md").read_text(encoding="utf-8")
    for index, chapter in enumerate(CHAPTERS, start=1):
        template = template.replace(
            f"{{{{CHAPTER_{index:02d}}}}}", chapter.read_text(encoding="utf-8").strip()
        )
    if "{{CHAPTER_" in template:
        raise RuntimeError("Unresolved chapter placeholder")
    OUT.mkdir(parents=True, exist_ok=True)
    target = OUT / "GRC_Risk_Register_and_Risk_Treatment_Practical_Manual_English_v1.0.md"
    target.write_text(template.strip() + "\n", encoding="utf-8")
    return target


def read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        raise RuntimeError(f"Empty CSV: {path}")
    return rows[0], rows[1:]


def style_sheet(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column in ws.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
        ws.column_dimensions[get_column_letter(column[0].column)].width = max(width, 12)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_workbook() -> Path:
    register_header, register_rows = read_csv(ROOT / "templates/Risk_Register_Template.csv")
    treatment_header, treatment_rows = read_csv(ROOT / "templates/Risk_Treatment_Plan_Template.csv")

    wb = Workbook()
    register = wb.active
    register.title = "Risk Register"
    register.append(register_header)
    for row in register_rows:
        register.append(row)
    style_sheet(register)

    treatment = wb.create_sheet("Treatment Plan")
    treatment.append(treatment_header)
    for row in treatment_rows:
        treatment.append(row)
    style_sheet(treatment)

    matrix = wb.create_sheet("Scoring Matrix")
    matrix.append(["Likelihood / Impact", 1, 2, 3, 4, 5])
    for likelihood in range(1, 6):
        matrix.append([likelihood] + [likelihood * impact for impact in range(1, 6)])
    matrix.append([])
    matrix.append(["Illustrative band", "Score range"])
    matrix.append(["Low", "1-4"])
    matrix.append(["Moderate", "5-9"])
    matrix.append(["High", "10-16"])
    matrix.append(["Critical", "17-25"])
    style_sheet(matrix, "B2")

    lists = wb.create_sheet("Validation Lists")
    values = {
        "Status": ["Draft", "Open", "Monitoring", "Accepted", "Closed"],
        "Response": ["Avoid", "Mitigate", "Transfer/Share", "Accept", "Pursue/Enhance"],
        "Trend": ["Improving", "Stable", "Deteriorating", "Unknown"],
        "Confidence": ["High", "Medium", "Low"],
        "Likelihood": [1, 2, 3, 4, 5],
        "Impact": [1, 2, 3, 4, 5],
    }
    for col, (name, entries) in enumerate(values.items(), start=1):
        lists.cell(1, col, name)
        for row, value in enumerate(entries, start=2):
            lists.cell(row, col, value)
    style_sheet(lists)

    instructions = wb.create_sheet("Instructions", 0)
    instructions.append(["GRC Risk Register and Risk Treatment Toolkit"])
    instructions.append(["Use", "Adapt the fields, scales, and approval rules to the organization’s approved methodology."])
    instructions.append(["Evidence", "Record current evidence, dates, assumptions, and accountable owners."])
    instructions.append(["Scoring", "The included 5x5 matrix is illustrative and must be formally approved before use."])
    instructions.append(["Acceptance", "Acceptance must be authorized, documented, monitored, and time-bound."])
    instructions.append(["Limitations", "This workbook does not establish compliance or replace professional advice."])
    instructions.column_dimensions["A"].width = 22
    instructions.column_dimensions["B"].width = 95
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    instructions["A1"].font = Font(bold=True, size=14)

    # Add bounded validation to common register columns when matching headers exist.
    header_index = {name: index + 1 for index, name in enumerate(register_header)}
    validation_map = {
        "Status": "'Validation Lists'!$A$2:$A$6",
        "Response": "'Validation Lists'!$B$2:$B$6",
        "Trend": "'Validation Lists'!$C$2:$C$5",
        "Confidence": "'Validation Lists'!$D$2:$D$4",
    }
    for header, formula in validation_map.items():
        if header in header_index:
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            register.add_data_validation(dv)
            col = get_column_letter(header_index[header])
            dv.add(f"{col}2:{col}1000")

    target = OUT / "GRC_Risk_Register_and_Risk_Treatment_Toolkit_English_v1.0.xlsx"
    wb.save(target)
    return target


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def write_manifest(paths: list[Path]) -> None:
    manifest = {
        "manual": "GRC Risk Register and Risk Treatment Practical Manual",
        "version": "1.0",
        "language": "English",
        "files": [{"name": p.name, "sha256": sha256(p), "bytes": p.stat().st_size} for p in paths],
    }
    (OUT / "MANIFEST.json").write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")


if __name__ == "__main__":
    master = assemble_master()
    workbook = build_workbook()
    write_manifest([master, workbook])
    print(master)
    print(workbook)
